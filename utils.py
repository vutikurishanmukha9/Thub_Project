import io
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def validate_biometric_data(data):
    """Validate biometric scanner data format"""
    try:
        if not isinstance(data, dict):
            return False
        
        required_fields = ['card_id', 'timestamp']
        for field in required_fields:
            if field not in data or not data[field]:
                return False
        
        # Validate timestamp format
        timestamp = data.get('timestamp')
        try:
            datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
        except ValueError:
            logging.error(f"Invalid timestamp format: {timestamp}")
            return False
        
        return True
    except Exception as e:
        logging.error(f"Data validation error: {e}")
        return False

def generate_excel_report(attendance_data, filters):
    """Generate Excel report from attendance data"""
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="004466", end_color="004466", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title and filters
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = "ADITYA ATTENDANCE REPORT"
        title_cell.font = Font(size=16, bold=True, color="004466")
        title_cell.alignment = Alignment(horizontal="center")
        
        # Add filter information
        row_num = 3
        filter_info = []
        if filters.get('session'):
            filter_info.append(f"Session: {filters['session']}")
        if filters.get('campus'):
            filter_info.append(f"Campus: {filters['campus']}")
        if filters.get('course'):
            filter_info.append(f"Course: {filters['course']}")
        if filters.get('date_from'):
            filter_info.append(f"From: {filters['date_from']}")
        if filters.get('date_to'):
            filter_info.append(f"To: {filters['date_to']}")
        
        if filter_info:
            ws[f'A{row_num}'] = "Filters Applied: " + " | ".join(filter_info)
            ws[f'A{row_num}'].font = Font(italic=True)
            row_num += 1
        
        # Add generation timestamp
        row_num += 1
        ws[f'A{row_num}'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{row_num}'].font = Font(italic=True, size=10)
        
        # Add headers
        row_num += 2
        headers = [
            "S.No.", "Student Name", "Roll Number", "Session", "Campus", 
            "Course", "Date", "Time", "Location", "Status"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data rows
        for index, (attendance_record, student) in enumerate(attendance_data, 1):
            row_num += 1
            data_row = [
                index,
                student.name,
                student.roll_number,
                student.session,
                student.campus,
                student.course,
                attendance_record.scan_date.strftime('%Y-%m-%d'),
                attendance_record.scan_time.strftime('%H:%M:%S'),
                attendance_record.location,
                attendance_record.status.title()
            ]
            
            for col_num, value in enumerate(data_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                
                # Apply conditional formatting for status
                if col_num == 10:  # Status column
                    if value.lower() == 'present':
                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                        cell.font = Font(color="155724")
                    elif value.lower() == 'late':
                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                        cell.font = Font(color="856404")
                    elif value.lower() == 'absent':
                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                        cell.font = Font(color="721C24")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(30, max(12, max_length + 2))
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary section
        row_num += 3
        ws[f'A{row_num}'] = "SUMMARY"
        ws[f'A{row_num}'].font = Font(bold=True, size=14, color="004466")
        
        row_num += 1
        total_records = len(attendance_data)
        present_count = sum(1 for record, _ in attendance_data if record.status.lower() == 'present')
        late_count = sum(1 for record, _ in attendance_data if record.status.lower() == 'late')
        
        summary_data = [
            ("Total Records:", total_records),
            ("Present:", present_count),
            ("Late:", late_count),
            ("Attendance Rate:", f"{(present_count/total_records*100):.1f}%" if total_records > 0 else "0%")
        ]
        
        for label, value in summary_data:
            ws[f'A{row_num}'] = label
            ws[f'B{row_num}'] = value
            ws[f'A{row_num}'].font = Font(bold=True)
            row_num += 1
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logging.info(f"Excel report generated with {total_records} records")
        return output
        
    except Exception as e:
        logging.error(f"Excel generation error: {e}")
        raise

def format_attendance_summary(attendance_data):
    """Format attendance data for display"""
    try:
        if not attendance_data:
            return {
                'total': 0,
                'present': 0,
                'late': 0,
                'absent': 0,
                'percentage': 0
            }
        
        total = len(attendance_data)
        present = sum(1 for record, _ in attendance_data if record.status.lower() == 'present')
        late = sum(1 for record, _ in attendance_data if record.status.lower() == 'late')
        absent = sum(1 for record, _ in attendance_data if record.status.lower() == 'absent')
        
        percentage = (present / total * 100) if total > 0 else 0
        
        return {
            'total': total,
            'present': present,
            'late': late,
            'absent': absent,
            'percentage': round(percentage, 1)
        }
        
    except Exception as e:
        logging.error(f"Summary formatting error: {e}")
        return {
            'total': 0,
            'present': 0,
            'late': 0,
            'absent': 0,
            'percentage': 0
        }

def sanitize_filename(filename):
    """Sanitize filename for safe file operations"""
    import re
    # Remove invalid characters
    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
    # Limit length
    if len(filename) > 100:
        name, ext = filename.rsplit('.', 1) if '.' in filename else (filename, '')
        filename = name[:95] + ('.' + ext if ext else '')
    
    return filename

def log_attendance_activity(activity_type, details):
    """Log attendance-related activities"""
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        logging.info(f"[{timestamp}] {activity_type}: {details}")
    except Exception as e:
        logging.error(f"Logging error: {e}")
